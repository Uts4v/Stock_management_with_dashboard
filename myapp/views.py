from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib import messages
from django.conf import settings
from django.db.models import Sum, Q
from django.utils import timezone
from django.http import HttpResponse
from datetime import timedelta, datetime, date
import traceback
import requests
from rest_framework import viewsets, status
from rest_framework.decorators import action, api_view
from rest_framework.response import Response
from rest_framework.views import APIView
from .models import Product, ProductVariant, StockTransaction
from .serializers import (
    ProductSerializer, ProductVariantSerializer,
    SellSerializer, RestockSerializer,
    RevenueReportSerializer, TopProductSerializer,
    ProductDictSerializer, ProductVariantDictSerializer,
)

# Supabase integration
from .supabase_client import (
    get_supabase_client,
    get_all_products, get_product_by_id,
    create_product, update_product, delete_product,
    get_all_variants, get_variant_by_id, get_variants_by_ids,
    create_variant, update_variant, delete_variant,
    get_all_transactions, create_transaction,
)

# Try to import openpyxl for Excel export
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# ─────────────────────────────────────────────
#  Helper: sync StockTransaction to Supabase
# ─────────────────────────────────────────────
def _sync_transaction(transaction):
    """Push a freshly-created StockTransaction row to Supabase."""
    try:
        create_transaction({
            'product_id': transaction.product_id,
            'variant_id': transaction.variant_id,
            'transaction_type': transaction.transaction_type,
            'quantity': transaction.quantity,
            'note': transaction.note or '',
            'created_by_id': transaction.created_by_id,
        })
    except Exception as e:
        print(f"Supabase sync error (transaction): {e}")


# ─────────────────────────────────────────────
#  Helper: sync variants from Supabase → SQLite
# ─────────────────────────────────────────────
def _sync_variants_from_supabase():
    """Pull all variants from Supabase and upsert into local SQLite."""
    try:
        supabase_variants = get_all_variants()
        for sv in supabase_variants:
            product_id = sv.get('product_id')
            if not product_id:
                continue
            if not Product.objects.filter(id=product_id).exists():
                continue
            ProductVariant.objects.update_or_create(
                id=sv['id'],
                defaults={
                    'product_id': product_id,
                    'variant_type': sv.get('variant_type', ''),
                    'variant_value': sv.get('variant_value', ''),
                    'stock': sv.get('stock', 0),
                    'min_stock': sv.get('min_stock', 0),
                    'cost_price': sv.get('cost_price') or 0,
                    'selling_price': sv.get('selling_price') or 0,
                    'barcode': sv.get('barcode') or '',
                }
            )
    except Exception as e:
        print(f"Supabase sync error (variants from Supabase): {e}")


# ─────────────────────────────────────────────
#  Auth / utility endpoints
# ─────────────────────────────────────────────
@api_view(['GET'])
def supabase_example(request):
    try:
        client = get_supabase_client()
        return Response({'status': 'success', 'message': 'Supabase connection successful', 'url': client.supabase_url})
    except Exception as e:
        return Response({'status': 'error', 'message': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
def current_user(request):
    if not request.user.is_authenticated:
        return Response({'error': 'Not authenticated'}, status=status.HTTP_401_UNAUTHORIZED)
    user = request.user
    return Response({
        'id': user.id, 'username': user.username, 'email': user.email,
        'first_name': user.first_name, 'last_name': user.last_name,
        'date_joined': user.date_joined.isoformat(), 'is_staff': user.is_staff,
    })


# ─────────────────────────────────────────────
#  ProductViewSet
# ─────────────────────────────────────────────
class ProductViewSet(viewsets.ModelViewSet):
    queryset = Product.objects.all()
    serializer_class = ProductSerializer

    def get_queryset(self):
        queryset = Product.objects.all()
        search = self.request.query_params.get('search', None)
        if search:
            queryset = queryset.filter(Q(name__icontains=search) | Q(category__icontains=search))
        category = self.request.query_params.get('category', None)
        if category:
            queryset = queryset.filter(category__iexact=category)
        return queryset

    def list(self, request, *args, **kwargs):
        """Always read from Supabase (single source of truth), fallback to SQLite."""
        try:
            products = get_all_products()
            search = request.query_params.get('search', None)
            category = request.query_params.get('category', None)
            if search:
                sl = search.lower()
                products = [p for p in products if sl in p.get('name', '').lower() or sl in p.get('category', '').lower()]
            if category:
                products = [p for p in products if p.get('category', '').lower() == category.lower()]
            serializer = ProductDictSerializer(products, many=True)
            return Response(serializer.data)
        except Exception as e:
            print(f"Supabase list error, falling back to SQLite: {e}")
            return super().list(request, *args, **kwargs)

    def perform_create(self, serializer):
        instance = serializer.save()
        try:
            create_product({
                'id': instance.id,
                'name': instance.name,
                'barcode': instance.barcode or '',
                'category': instance.category or '',
                'stock': instance.stock,
                'min_stock': instance.min_stock,
                'cost_price': float(instance.cost_price),
                'selling_price': float(instance.selling_price),
                'created_at': instance.created_at.isoformat(),
            })
        except Exception as e:
            print(f"Supabase sync error (create product): {e}")

    def perform_update(self, serializer):
        instance = serializer.save()
        try:
            update_product(instance.id, {
                'name': instance.name,
                'barcode': instance.barcode or '',
                'category': instance.category or '',
                'stock': instance.stock,
                'min_stock': instance.min_stock,
                'cost_price': float(instance.cost_price),
                'selling_price': float(instance.selling_price),
            })
        except Exception as e:
            print(f"Supabase sync error (update product): {e}")

    def destroy(self, request, *args, **kwargs):
        pk = kwargs.get('pk')
        try:
            instance = self.get_object()
            instance.delete()
        except Exception:
            pass
        try:
            delete_product(int(pk))
        except Exception as e:
            print(f"Supabase sync error (delete product): {e}")
        return Response(status=status.HTTP_204_NO_CONTENT)

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)

    # ── sell ──────────────────────────────────
    @action(detail=True, methods=['post'])
    def sell(self, request, pk=None):
        try:
            serializer = SellSerializer(data=request.data)
            if not serializer.is_valid():
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
            quantity = serializer.validated_data['quantity']
            note = serializer.validated_data.get('note', '')

            try:
                product = Product.objects.get(pk=pk)
            except Product.DoesNotExist:
                return Response({'error': 'Product not found'}, status=status.HTTP_404_NOT_FOUND)

            if product.stock < quantity:
                return Response({'error': f'Insufficient stock. Available: {product.stock}'}, status=status.HTTP_400_BAD_REQUEST)

            product.stock -= quantity
            product.save()

            txn = StockTransaction.objects.create(
                product=product,
                transaction_type=StockTransaction.TransactionType.SALE,
                quantity=quantity, note=note,
                created_by=request.user if request.user.is_authenticated else None
            )

            try:
                update_product(product.id, {'stock': product.stock})
            except Exception as e:
                print(f"Supabase sync error (sell product stock): {e}")

            try:
                _sync_transaction(txn)
            except Exception as e:
                print(f"Supabase sync error (txn): {e}")

            return Response({'message': f'Sold {quantity} units of {product.name}', 'product': ProductSerializer(product).data})

        except Exception as e:
            print(f"SELL ERROR: {traceback.format_exc()}")
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    # ── restock ───────────────────────────────
    @action(detail=True, methods=['post'])
    def restock(self, request, pk=None):
        try:
            serializer = RestockSerializer(data=request.data)
            if not serializer.is_valid():
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
            quantity = serializer.validated_data['quantity']
            note = serializer.validated_data.get('note', '')

            try:
                product = Product.objects.get(pk=pk)
            except Product.DoesNotExist:
                return Response({'error': 'Product not found'}, status=status.HTTP_404_NOT_FOUND)

            product.stock += quantity
            product.save()

            txn = StockTransaction.objects.create(
                product=product,
                transaction_type=StockTransaction.TransactionType.RESTOCK,
                quantity=quantity, note=note,
                created_by=request.user if request.user.is_authenticated else None
            )

            try:
                update_product(product.id, {'stock': product.stock})
            except Exception as e:
                print(f"Supabase sync error (restock product stock): {e}")

            try:
                _sync_transaction(txn)
            except Exception as e:
                print(f"Supabase sync error (txn): {e}")

            return Response({'message': f'Restocked {quantity} units of {product.name}', 'product': ProductSerializer(product).data})

        except Exception as e:
            print(f"RESTOCK ERROR: {traceback.format_exc()}")
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    # ── update_stock ──────────────────────────
    @action(detail=True, methods=['patch'], url_path='update-stock')
    def update_stock(self, request, pk=None):
        new_stock = request.data.get('stock')
        if new_stock is None:
            return Response({'error': 'stock is required'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            new_stock = int(new_stock)
            if new_stock < 0:
                raise ValueError
        except (ValueError, TypeError):
            return Response({'error': 'stock must be a non-negative integer'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            product = Product.objects.get(pk=pk)
        except Product.DoesNotExist:
            return Response({'error': 'Product not found'}, status=status.HTTP_404_NOT_FOUND)

        product.stock = new_stock
        product.save()
        try:
            update_product(product.id, {'stock': new_stock})
        except Exception as e:
            print(f"Supabase sync error (update stock): {e}")
        return Response({'message': 'Stock updated', 'stock': new_stock})

    # ── update_min_stock ──────────────────────
    @action(detail=True, methods=['patch'])
    def update_min_stock(self, request, pk=None):
        min_stock = request.data.get('min_stock')
        if min_stock is None:
            return Response({'error': 'min_stock is required'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            min_stock = int(min_stock)
            if min_stock < 0:
                raise ValueError
        except (ValueError, TypeError):
            return Response({'error': 'min_stock must be a non-negative integer'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            product = Product.objects.get(pk=pk)
        except Product.DoesNotExist:
            return Response({'error': 'Product not found'}, status=status.HTTP_404_NOT_FOUND)

        product.min_stock = min_stock
        product.save()
        try:
            update_product(product.id, {'min_stock': min_stock})
        except Exception as e:
            print(f"Supabase sync error (update min_stock): {e}")
        return Response({'message': 'Min stock updated', 'min_stock': min_stock})


# ─────────────────────────────────────────────
#  ProductVariantViewSet
# ─────────────────────────────────────────────
class ProductVariantViewSet(viewsets.ModelViewSet):
    queryset = ProductVariant.objects.all()
    serializer_class = ProductVariantSerializer

    def get_queryset(self):
        queryset = ProductVariant.objects.select_related('product').all()
        product_id = self.request.query_params.get('product_id', None)
        if product_id:
            queryset = queryset.filter(product_id=product_id)
        return queryset

    def list(self, request, *args, **kwargs):
        """Read variants from Supabase, fallback to SQLite."""
        try:
            variants = get_all_variants()
            product_id = request.query_params.get('product_id', None)
            if product_id:
                variants = [v for v in variants if v.get('product_id') == int(product_id)]
            serializer = ProductVariantDictSerializer(variants, many=True)
            return Response(serializer.data)
        except Exception as e:
            print(f"Supabase list variants error, falling back to SQLite: {e}")
            return super().list(request, *args, **kwargs)

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        variant = serializer.save()
        variant.product.update_total_stock()

        try:
            create_variant({
                'id': variant.id,
                'product_id': variant.product_id,
                'variant_type': variant.variant_type,
                'variant_value': variant.variant_value,
                'stock': variant.stock,
                'min_stock': variant.min_stock,
                'cost_price': float(variant.cost_price) if variant.cost_price else None,
                'selling_price': float(variant.selling_price) if variant.selling_price else None,
                'barcode': variant.barcode or '',
                'created_at': variant.created_at.isoformat(),
            })
        except Exception as e:
            print(f"Supabase sync error (create variant): {e}")

        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        variant = serializer.save()
        variant.product.update_total_stock()

        try:
            update_variant(variant.id, {
                'product_id': variant.product_id,
                'variant_type': variant.variant_type,
                'variant_value': variant.variant_value,
                'stock': variant.stock,
                'min_stock': variant.min_stock,
                'cost_price': float(variant.cost_price) if variant.cost_price else None,
                'selling_price': float(variant.selling_price) if variant.selling_price else None,
                'barcode': variant.barcode or '',
            })
        except Exception as e:
            print(f"Supabase sync error (update variant): {e}")

        return Response(serializer.data)

    def destroy(self, request, *args, **kwargs):
        pk = kwargs.get('pk')
        try:
            instance = self.get_object()
            product = instance.product
            instance.delete()
            product.update_total_stock()
        except Exception:
            pass
        try:
            delete_variant(int(pk))
        except Exception as e:
            print(f"Supabase sync error (delete variant): {e}")
        return Response(status=status.HTTP_204_NO_CONTENT)

    # ── sell ──────────────────────────────────
    @action(detail=True, methods=['post'])
    def sell(self, request, pk=None):
        try:
            serializer = SellSerializer(data=request.data)
            if not serializer.is_valid():
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
            quantity = serializer.validated_data['quantity']
            note = serializer.validated_data.get('note', '')

            try:
                variant = ProductVariant.objects.get(pk=pk)
            except ProductVariant.DoesNotExist:
                return Response({'error': 'Variant not found'}, status=status.HTTP_404_NOT_FOUND)

            if variant.stock < quantity:
                return Response({'error': f'Insufficient stock. Available: {variant.stock}'}, status=status.HTTP_400_BAD_REQUEST)

            variant.stock -= quantity
            variant.save()
            variant.product.update_total_stock()

            txn = StockTransaction.objects.create(
                product=variant.product, variant=variant,
                transaction_type=StockTransaction.TransactionType.SALE,
                quantity=quantity, note=note,
                created_by=request.user if request.user.is_authenticated else None
            )

            try:
                update_variant(variant.id, {'stock': variant.stock})
                update_product(variant.product_id, {'stock': variant.product.stock})
            except Exception as e:
                print(f"Supabase sync error (sell variant): {e}")

            try:
                _sync_transaction(txn)
            except Exception as e:
                print(f"Supabase sync error (txn): {e}")

            return Response({
                'message': f'Sold {quantity} units of {variant.product.name} ({variant.variant_type}: {variant.variant_value})',
                'variant': ProductVariantSerializer(variant).data
            })

        except Exception as e:
            print(f"VARIANT SELL ERROR: {traceback.format_exc()}")
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    # ── restock ───────────────────────────────
    @action(detail=True, methods=['post'])
    def restock(self, request, pk=None):
        try:
            serializer = RestockSerializer(data=request.data)
            if not serializer.is_valid():
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
            quantity = serializer.validated_data['quantity']
            note = serializer.validated_data.get('note', '')

            try:
                variant = ProductVariant.objects.get(pk=pk)
            except ProductVariant.DoesNotExist:
                return Response({'error': 'Variant not found'}, status=status.HTTP_404_NOT_FOUND)

            variant.stock += quantity
            variant.save()
            variant.product.update_total_stock()

            txn = StockTransaction.objects.create(
                product=variant.product, variant=variant,
                transaction_type=StockTransaction.TransactionType.RESTOCK,
                quantity=quantity, note=note,
                created_by=request.user if request.user.is_authenticated else None
            )

            try:
                update_variant(variant.id, {'stock': variant.stock})
                update_product(variant.product_id, {'stock': variant.product.stock})
            except Exception as e:
                print(f"Supabase sync error (restock variant): {e}")

            try:
                _sync_transaction(txn)
            except Exception as e:
                print(f"Supabase sync error (txn): {e}")

            return Response({
                'message': f'Restocked {quantity} units of {variant.product.name} ({variant.variant_type}: {variant.variant_value})',
                'variant': ProductVariantSerializer(variant).data
            })

        except Exception as e:
            print(f"VARIANT RESTOCK ERROR: {traceback.format_exc()}")
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    # ── update_stock ──────────────────────────
    @action(detail=True, methods=['patch'], url_path='update-stock')
    def update_stock(self, request, pk=None):
        new_stock = request.data.get('stock')
        if new_stock is None:
            return Response({'error': 'stock is required'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            new_stock = int(new_stock)
            if new_stock < 0:
                raise ValueError
        except (ValueError, TypeError):
            return Response({'error': 'stock must be a non-negative integer'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            variant = ProductVariant.objects.get(pk=pk)
        except ProductVariant.DoesNotExist:
            return Response({'error': 'Variant not found'}, status=status.HTTP_404_NOT_FOUND)

        variant.stock = new_stock
        variant.save()
        variant.product.update_total_stock()

        try:
            update_variant(variant.id, {'stock': new_stock})
            update_product(variant.product_id, {'stock': variant.product.stock})
        except Exception as e:
            print(f"Supabase sync error (update variant stock): {e}")
        return Response({'message': 'Stock updated', 'stock': new_stock})

    # ── update_min_stock ──────────────────────
    @action(detail=True, methods=['patch'])
    def update_min_stock(self, request, pk=None):
        min_stock = request.data.get('min_stock')
        if min_stock is None:
            return Response({'error': 'min_stock is required'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            min_stock = int(min_stock)
            if min_stock < 0:
                raise ValueError
        except (ValueError, TypeError):
            return Response({'error': 'min_stock must be a non-negative integer'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            variant = ProductVariant.objects.get(pk=pk)
        except ProductVariant.DoesNotExist:
            return Response({'error': 'Variant not found'}, status=status.HTTP_404_NOT_FOUND)

        variant.min_stock = min_stock
        variant.save()
        try:
            update_variant(variant.id, {'min_stock': min_stock})
        except Exception as e:
            print(f"Supabase sync error (update variant min_stock): {e}")
        return Response({'message': 'Min stock updated', 'min_stock': min_stock})


# ─────────────────────────────────────────────
#  Categories
# ─────────────────────────────────────────────
class CategoriesView(APIView):
    def get(self, request):
        try:
            products = get_all_products()
            categories = sorted(set(p.get('category', '') for p in products if p.get('category')))
            return Response(categories)
        except Exception:
            pass
        categories = Product.objects.values_list('category', flat=True).distinct().order_by('category')
        return Response(list(categories))


# ─────────────────────────────────────────────
#  Sales History
# ─────────────────────────────────────────────
class SalesHistoryView(APIView):
    def get(self, request):
        trans_type = request.query_params.get('type', None)
        start_date = request.query_params.get('start_date', None)
        end_date   = request.query_params.get('end_date', None)
        product_id = request.query_params.get('product_id', None)

        try:
            client = get_supabase_client()
            query = client.table('myapp_stocktransaction').select(
                "*, myapp_product(name, selling_price)"
            ).order('created_at', desc=True)

            if trans_type:
                query = query.eq('transaction_type', trans_type)
            if start_date:
                query = query.gte('created_at', f'{start_date}T00:00:00')
            if end_date:
                query = query.lte('created_at', f'{end_date}T23:59:59')
            if product_id:
                query = query.eq('product_id', int(product_id))

            result = query.limit(500).execute()
            raw = result.data or []

            # Bulk fetch variants to avoid N+1 queries
            variant_ids = list({t.get('variant_id') for t in raw if t.get('variant_id')})
            variants_lookup = {}
            if variant_ids:
                try:
                    variants_data = get_variants_by_ids(variant_ids)
                    for v in variants_data:
                        variants_lookup[v.get('id')] = v
                except Exception as e:
                    print(f"Error bulk fetching variants: {e}")

            transactions = []
            for t in raw:
                product_data  = t.get('myapp_product') or {}
                product_name  = product_data.get('name', 'Unknown')
                selling_price = float(product_data.get('selling_price') or 0)

                variant_id    = t.get('variant_id')
                variant_label = None
                if variant_id:
                    v = variants_lookup.get(variant_id)
                    if not v:
                        try:
                            v = get_variant_by_id(variant_id)
                            if v:
                                variants_lookup[variant_id] = v
                        except Exception:
                            pass
                    if v:
                        variant_label = f"{v.get('variant_type')}: {v.get('variant_value')}"
                        if v.get('selling_price'):
                            selling_price = float(v['selling_price'])

                revenue = (
                    round(float(t.get('quantity', 0)) * selling_price, 2)
                    if t.get('transaction_type') == 'SALE' else None
                )

                transactions.append({
                    'id':               t.get('id'),
                    'product_id':       t.get('product_id'),
                    'product_name':     product_name,
                    'variant_id':       variant_id,
                    'variant_label':    variant_label,
                    'transaction_type': t.get('transaction_type'),
                    'quantity':         t.get('quantity'),
                    'unit_price':       selling_price,
                    'revenue':          revenue,
                    'note':             t.get('note') or '',
                    'created_at':       t.get('created_at', ''),
                })

            return Response(transactions)

        except Exception as e:
            print(f"Supabase transactions error, falling back to SQLite: {e}")
            queryset = StockTransaction.objects.select_related('product', 'variant').all()
            if trans_type:
                queryset = queryset.filter(transaction_type=trans_type)
            if start_date:
                queryset = queryset.filter(created_at__date__gte=start_date)
            if end_date:
                queryset = queryset.filter(created_at__date__lte=end_date)
            if product_id:
                queryset = queryset.filter(product_id=product_id)
            queryset = queryset.order_by('-created_at')

            transactions = []
            for t in queryset[:500]:
                unit_price = (
                    float(t.variant.selling_price) if t.variant and t.variant.selling_price
                    else float(t.product.selling_price)
                )
                revenue = (
                    round(float(t.quantity) * unit_price, 2)
                    if t.transaction_type == StockTransaction.TransactionType.SALE else None
                )
                transactions.append({
                    'id': t.id,
                    'product_id': t.product.id,
                    'product_name': t.product.name,
                    'variant_id': t.variant_id,
                    'variant_label': f"{t.variant.variant_type}: {t.variant.variant_value}" if t.variant else None,
                    'transaction_type': t.transaction_type,
                    'quantity': t.quantity,
                    'unit_price': unit_price,
                    'revenue': revenue,
                    'note': t.note or '',
                    'created_at': t.created_at.isoformat(),
                })
            return Response(transactions)


# ─────────────────────────────────────────────
#  Sales Summary
# ─────────────────────────────────────────────
class SalesSummaryView(APIView):
    def get(self, request):
        start_date = request.query_params.get('start_date', None)
        end_date   = request.query_params.get('end_date', None)

        try:
            client = get_supabase_client()
            query = client.table('myapp_stocktransaction').select(
                "quantity, myapp_product(selling_price)"
            ).eq('transaction_type', 'SALE')

            if start_date:
                query = query.gte('created_at', f'{start_date}T00:00:00')
            if end_date:
                query = query.lte('created_at', f'{end_date}T23:59:59')

            result = query.execute()
            sales = result.data or []

            total_sales    = len(sales)
            total_quantity = sum(t.get('quantity', 0) for t in sales)
            total_revenue  = sum(
                float(t.get('quantity', 0)) * float((t.get('myapp_product') or {}).get('selling_price') or 0)
                for t in sales
            )
            avg_sale_value = total_revenue / total_sales if total_sales > 0 else 0

            return Response({
                'total_transactions':  total_sales,
                'total_quantity_sold': total_quantity,
                'total_revenue':       round(total_revenue, 2),
                'average_sale_value':  round(avg_sale_value, 2),
            })

        except Exception as e:
            print(f"Supabase summary error, falling back to SQLite: {e}")
            queryset = StockTransaction.objects.filter(
                transaction_type=StockTransaction.TransactionType.SALE
            ).select_related('product', 'variant')
            if start_date:
                queryset = queryset.filter(created_at__date__gte=start_date)
            if end_date:
                queryset = queryset.filter(created_at__date__lte=end_date)

            total_sales    = queryset.count()
            total_quantity = queryset.aggregate(Sum('quantity'))['quantity__sum'] or 0
            total_revenue  = sum(
                float(t.quantity) * float(
                    t.variant.selling_price if t.variant and t.variant.selling_price
                    else t.product.selling_price
                )
                for t in queryset
            )
            avg_sale_value = total_revenue / total_sales if total_sales > 0 else 0

            return Response({
                'total_transactions':  total_sales,
                'total_quantity_sold': total_quantity,
                'total_revenue':       round(total_revenue, 2),
                'average_sale_value':  round(avg_sale_value, 2),
            })


# ─────────────────────────────────────────────
#  Revenue Report
# ─────────────────────────────────────────────
class RevenueReportView(APIView):
    def get(self, request):
        period = request.query_params.get('period', 'daily')
        try:
            client = get_supabase_client()
            result = client.table('myapp_stocktransaction').select(
                "quantity, created_at, myapp_product(selling_price)"
            ).eq('transaction_type', 'SALE').execute()
            sales = result.data or []

            grouped = {}
            for t in sales:
                unit_price = float((t.get('myapp_product') or {}).get('selling_price') or 0)
                revenue    = float(t.get('quantity', 0)) * unit_price
                raw_date   = t.get('created_at', '')[:10]
                try:
                    d = date.fromisoformat(raw_date)
                except Exception:
                    continue

                if period == 'weekly':
                    key = d - timedelta(days=d.weekday())
                elif period == 'monthly':
                    key = d.replace(day=1)
                else:
                    key = d

                if key not in grouped:
                    grouped[key] = {'total_revenue': 0.0, 'total_quantity': 0}
                grouped[key]['total_revenue']  += revenue
                grouped[key]['total_quantity'] += t.get('quantity', 0)

            return Response([
                {'date': str(k), 'total_revenue': round(v['total_revenue'], 2), 'total_quantity': v['total_quantity']}
                for k, v in sorted(grouped.items())
            ])
        except Exception as e:
            print(f"Revenue report error: {e}")
            return Response([])


# ─────────────────────────────────────────────
#  Top Products
# ─────────────────────────────────────────────
class TopProductsView(APIView):
    def get(self, request):
        today         = timezone.now().date()
        start_of_week = today - timedelta(days=today.weekday())
        search        = request.query_params.get('search', None)

        try:
            client = get_supabase_client()
            result = client.table('myapp_stocktransaction').select(
                "quantity, product_id, myapp_product(name, selling_price)"
            ).eq('transaction_type', 'SALE').gte(
                'created_at', f'{start_of_week}T00:00:00'
            ).lte('created_at', f'{today}T23:59:59').execute()
            sales = result.data or []
        except Exception as e:
            print(f"Top products Supabase error: {e}")
            sales = []

        product_stats = {}
        for sale in sales:
            pid        = sale.get('product_id')
            pdata      = sale.get('myapp_product') or {}
            unit_price = float(pdata.get('selling_price') or 0)
            if pid not in product_stats:
                product_stats[pid] = {
                    'product_id': pid,
                    'product_name': pdata.get('name', 'Unknown'),
                    'total_units_sold': 0,
                    'total_revenue': 0.0,
                }
            product_stats[pid]['total_units_sold'] += sale.get('quantity', 0)
            product_stats[pid]['total_revenue']    += float(sale.get('quantity', 0)) * unit_price

        sorted_products = sorted(product_stats.values(), key=lambda x: x['total_units_sold'], reverse=True)[:5]
        for p in sorted_products:
            p['total_revenue'] = round(p['total_revenue'], 2)

        if search:
            sl       = search.lower()
            filtered = [p for p in sorted_products if sl in p['product_name'].lower()]
            if not filtered:
                return Response({'message': f'"{search}" not found in top 5 products this week',
                                 'search_term': search, 'top_products': sorted_products})
            return Response({'search_term': search, 'highlighted_products': filtered, 'top_products': sorted_products})

        return Response(sorted_products)


# ─────────────────────────────────────────────
#  Delete Transactions
# ─────────────────────────────────────────────
@api_view(['DELETE'])
def delete_transactions(request):
    clear_all = request.data.get('clear_all', False)

    if clear_all:
        count = StockTransaction.objects.count()
        StockTransaction.objects.all().delete()
        try:
            client = get_supabase_client()
            client.table('myapp_stocktransaction').delete().neq('id', 0).execute()
        except Exception as e:
            print(f"Supabase delete all transactions error: {e}")
        return Response({'message': f'Deleted {count} transactions', 'deleted_count': count})

    ids = request.data.get('ids', [])
    if not ids:
        return Response({'error': 'No transaction IDs provided'}, status=status.HTTP_400_BAD_REQUEST)

    deleted_count = StockTransaction.objects.filter(id__in=ids).delete()[0]
    try:
        client = get_supabase_client()
        for tid in ids:
            client.table('myapp_stocktransaction').delete().eq('id', tid).execute()
    except Exception as e:
        print(f"Supabase delete transactions error: {e}")
    return Response({'message': f'Deleted {deleted_count} transactions', 'deleted_count': deleted_count})


# ─────────────────────────────────────────────
#  Excel Export – Products
# ─────────────────────────────────────────────
@api_view(['GET'])
def export_products_excel(request):
    if not OPENPYXL_AVAILABLE:
        return Response({'error': 'openpyxl not installed. Run: pip install openpyxl'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    products = Product.objects.all().order_by('-created_at')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Products"

    hf    = Font(bold=True, color="FFFFFF")
    hfill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for col, h in enumerate(['S.N','Product Name','Barcode','Category','Stock','Min Stock','Cost Price','Selling Price','Status'], 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hf; cell.fill = hfill; cell.alignment = Alignment(horizontal='center'); cell.border = border

    for row, p in enumerate(products, 2):
        for col, val in enumerate([row-1, p.name, p.barcode or '', p.category or '', p.stock, p.min_stock,
                                    float(p.cost_price), float(p.selling_price), p.status], 1):
            ws.cell(row=row, column=col, value=val).border = border

    for i, w in enumerate([5,30,15,15,10,10,12,12,10], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Stock_Management_{datetime.now().strftime("%Y%m%d")}.xlsx'
    wb.save(response)
    return response


# ─────────────────────────────────────────────
#  Excel Export – Transactions
# ─────────────────────────────────────────────
@api_view(['POST'])
def export_transactions_excel(request):
    if not OPENPYXL_AVAILABLE:
        return Response({'error': 'openpyxl not installed.'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    transactions = request.data.get('transactions', [])
    if not transactions:
        return Response({'error': 'No transactions to export'}, status=status.HTTP_400_BAD_REQUEST)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales History"

    hf    = Font(bold=True, color="FFFFFF")
    hfill = PatternFill(start_color="28a745", end_color="28a745", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for col, h in enumerate(['S.N','Date','Time','Product','Type','Quantity','Unit Price','Total Value','Note'], 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hf; cell.fill = hfill; cell.alignment = Alignment(horizontal='center'); cell.border = border

    for row, t in enumerate(transactions, 2):
        dt = datetime.fromisoformat(t['created_at'].replace('Z', '+00:00'))
        tv = t.get('revenue')
        for col, val in enumerate([row-1, dt.strftime('%Y-%m-%d'), dt.strftime('%H:%M'),
                                    t['product_name'], t['transaction_type'], t['quantity'],
                                    float(t['unit_price']), float(tv) if tv is not None else '',
                                    t.get('note') or ''], 1):
            ws.cell(row=row, column=col, value=val).border = border

    for i, w in enumerate([5,12,10,30,12,10,12,12,25], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Sales_History_{datetime.now().strftime("%Y%m%d")}.xlsx'
    wb.save(response)
    return response


# ─────────────────────────────────────────────
#  Google Sheets – Export
# ─────────────────────────────────────────────
@api_view(['POST'])
def sync_google_sheet(request):
    sheet_id = request.data.get('sheet_id')
    api_key  = request.data.get('api_key')
    if not sheet_id or not api_key:
        return Response({'error': 'sheet_id and api_key are required'}, status=status.HTTP_400_BAD_REQUEST)
    try:
        products = Product.objects.all().order_by('-created_at')
        values   = [['Name','Barcode','Category','Stock','Min Stock','Cost Price','Selling Price','Status','Last Updated']]
        for p in products:
            values.append([p.name, p.barcode or '', p.category or '', p.stock, p.min_stock,
                           float(p.cost_price), float(p.selling_price), p.status, datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
        range_name = f'Sheet1!A1:I{len(values)}'
        resp = requests.put(
            f'https://sheets.googleapis.com/v4/spreadsheets/{sheet_id}/values/{range_name}',
            params={'key': api_key, 'valueInputOption': 'USER_ENTERED'}, json={'values': values}
        )
        if resp.status_code == 200:
            return Response({'message': f'Successfully synced {len(products)} products to Google Sheets', 'products_count': len(products)})
        error_msg = (resp.json() if resp.text else {}).get('error', {}).get('message', resp.text)
        if 'PERMISSION_DENIED' in str(error_msg) or resp.status_code == 403:
            return Response({'error': 'Permission denied. Share your Google Sheet with "Anyone with the link can edit".'}, status=status.HTTP_400_BAD_REQUEST)
        return Response({'error': f'Google Sheets API error: {error_msg}'}, status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# ─────────────────────────────────────────────
#  Google Sheets – Import
# ─────────────────────────────────────────────
@api_view(['POST'])
def import_from_google_sheet(request):
    sheet_id   = request.data.get('sheet_id')
    api_key    = request.data.get('api_key')
    range_name = request.data.get('range', 'Sheet1!A1:I100')
    if not sheet_id or not api_key:
        return Response({'error': 'sheet_id and api_key are required'}, status=status.HTTP_400_BAD_REQUEST)
    try:
        resp = requests.get(f'https://sheets.googleapis.com/v4/spreadsheets/{sheet_id}/values/{range_name}', params={'key': api_key})
        if resp.status_code != 200:
            return Response({'error': f'Google Sheets API error: {resp.text}'}, status=status.HTTP_400_BAD_REQUEST)
        values = resp.json().get('values', [])
        if not values or len(values) < 2:
            return Response({'error': 'No data found in the specified range'}, status=status.HTTP_400_BAD_REQUEST)

        imported_count = 0
        errors = []
        for i, row in enumerate(values[1:], 2):
            try:
                if len(row) < 6:
                    errors.append(f'Row {i}: Not enough columns'); continue
                name = row[0] if row else ''
                if not name:
                    errors.append(f'Row {i}: Product name is required'); continue
                barcode       = row[1] if len(row) > 1 else ''
                category      = row[2] if len(row) > 2 else ''
                stock         = int(row[3])   if len(row) > 3 and row[3] else 0
                min_stock     = int(row[4])   if len(row) > 4 and row[4] else 10
                cost_price    = float(row[5]) if len(row) > 5 and row[5] else 0
                selling_price = float(row[6]) if len(row) > 6 and row[6] else 0

                product, created = Product.objects.get_or_create(
                    name=name,
                    defaults={'barcode': barcode, 'category': category, 'stock': stock,
                              'min_stock': min_stock, 'cost_price': cost_price, 'selling_price': selling_price}
                )
                if not created:
                    product.barcode = barcode; product.category = category; product.stock = stock
                    product.min_stock = min_stock; product.cost_price = cost_price; product.selling_price = selling_price
                    product.save()

                try:
                    if created:
                        create_product({'id': product.id, 'name': product.name, 'barcode': product.barcode or '',
                                        'category': product.category or '', 'stock': product.stock,
                                        'min_stock': product.min_stock, 'cost_price': float(product.cost_price),
                                        'selling_price': float(product.selling_price), 'created_at': product.created_at.isoformat()})
                    else:
                        update_product(product.id, {'barcode': product.barcode or '', 'category': product.category or '',
                                                     'stock': product.stock, 'min_stock': product.min_stock,
                                                     'cost_price': float(product.cost_price), 'selling_price': float(product.selling_price)})
                except Exception as sup_err:
                    print(f"Supabase sync error (import row {i}): {sup_err}")

                imported_count += 1
            except Exception as e:
                errors.append(f'Row {i}: {str(e)}')

        return Response({'message': f'Successfully imported {imported_count} products',
                         'imported_count': imported_count, 'errors': errors if errors else None})
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# ─────────────────────────────────────────────
#  Template views
# ─────────────────────────────────────────────
@login_required
def index(request):
    return render(request, 'index.html')

@login_required
def low_stock(request):
    return render(request, 'low-stock.html')

@login_required
def sales_history(request):
    return render(request, 'sales-history.html')


# ─────────────────────────────────────────────
#  Auth views
# ─────────────────────────────────────────────
def user_login(request):
    if request.user.is_authenticated:
        return redirect('index')
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '')
        if not username or not password:
            messages.error(request, 'Please provide both username and password.')
            return render(request, 'login.html')
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            messages.success(request, f'Welcome back, {user.username}!')
            return redirect(request.GET.get('next', 'index'))
        messages.error(request, 'Invalid username or password.')
    return render(request, 'login.html')


def user_signup(request):
    if request.user.is_authenticated:
        return redirect('index')
    if request.method == 'POST':
        username         = request.POST.get('username', '').strip()
        email            = request.POST.get('email', '').strip()
        password         = request.POST.get('password', '').strip()
        password_confirm = request.POST.get('confirm_password', '').strip()
        if not username or not email or not password:
            messages.error(request, 'All fields are required.')
            return render(request, 'signup.html')
        if password != password_confirm:
            messages.error(request, 'Passwords do not match.')
            return render(request, 'signup.html')
        if len(password) < 8:
            messages.error(request, 'Password must be at least 8 characters.')
            return render(request, 'signup.html')
        if User.objects.filter(username=username).exists():
            messages.error(request, 'Username already exists.')
            return render(request, 'signup.html')
        if User.objects.filter(email=email).exists():
            messages.error(request, 'Email already registered.')
            return render(request, 'signup.html')
        User.objects.create_user(username=username, email=email, password=password)
        messages.success(request, 'Account created successfully! Please login.')
        return redirect('login')
    return render(request, 'signup.html')


def user_logout(request):
    logout(request)
    messages.success(request, 'You have been logged out.')
    return redirect('login')